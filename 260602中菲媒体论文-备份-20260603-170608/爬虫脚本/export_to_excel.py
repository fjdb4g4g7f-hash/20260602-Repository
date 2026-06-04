"""
Export scraped data to Excel format compatible with existing corpus

Usage: python3 export_to_excel.py
Output: pia_WPS_data.xlsx + dfa_WPS_data.xlsx  (if both exist)
        combined_WPS_data.xlsx  (merged)
"""

import json
import os
from datetime import datetime

# Try to import openpyxl, install if not available
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system('pip3 install openpyxl -q')
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

PROJECT_DIR = '/Users/keying/Desktop/Claude Code/260602中菲媒体论文'

def json_to_excel(json_file, output_file, source_name):
    """Convert a JSON scraper output to Excel format matching the existing corpus."""
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    articles = data.get('articles', [])
    if not articles:
        print(f"No articles found in {json_file}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = f"{source_name} Data"

    # Headers matching existing corpus (大表.xlsx)
    headers = ['Title', 'Date', 'Source', 'Country', 'Content', 'URL', 'In_Date_Range']
    from openpyxl.styles import Font
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Add data rows
    row = 2
    for article in articles:
        title = article.get('title', '')
        date = article.get('date', '')
        source = article.get('source', source_name)
        country = article.get('country', 'Philippines')
        content = article.get('content', '')
        url = article.get('url', '')
        in_range = article.get('inDateRange', True)

        # Normalize date format
        if date:
            try:
                dt = datetime.fromisoformat(date.replace('Z', '+00:00'))
                date = dt.strftime('%Y-%m-%d')
            except:
                pass

        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=2, value=date)
        ws.cell(row=row, column=3, value=source)
        ws.cell(row=row, column=4, value=country)
        ws.cell(row=row, column=5, value=content)
        ws.cell(row=row, column=6, value=url)
        ws.cell(row=row, column=7, value='Yes' if in_range else 'No')
        row += 1

    # Auto-adjust column widths (limit to reasonable values)
    for col in range(1, len(headers) + 1):
        max_length = min(
            max(len(str(ws.cell(row=r, column=col).value or '')) for r in range(1, row)) + 2,
            80  # cap at 80 chars wide
        )
        ws.column_dimensions[get_column_letter(col)].width = max_length

    wb.save(output_file)
    print(f"Saved {row - 2} rows to {output_file}")

    # Summary stats
    year_dist = {}
    for article in articles:
        d = article.get('date', '')
        if d and len(d) >= 4:
            y = d[:4]
            year_dist[y] = year_dist.get(y, 0) + 1

    print(f"Year distribution: {dict(sorted(year_dist.items()))}")

def merge_excels(file_list, output_file):
    """Merge multiple Excel files into one."""
    import shutil
    from openpyxl import load_workbook

    if not file_list:
        print("No files to merge")
        return

    # Start with first file
    merged_wb = Workbook()
    merged_ws = merged_wb.active
    merged_ws.title = "Combined WPS Data"

    headers_written = False
    total_rows = 0

    for file_path in file_list:
        if not os.path.exists(file_path):
            print(f"  Skipping {file_path} (not found)")
            continue

        wb = load_workbook(file_path)
        ws = wb.active

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if not headers_written:
                merged_ws.append(row)
                headers_written = True
            elif row_idx > 1:  # Skip header row
                merged_ws.append(row)
                total_rows += 1

        wb.close()

    merged_wb.save(output_file)
    print(f"\nMerged {total_rows} rows into {output_file}")

def main():
    # Export PIA data
    pia_json = os.path.join(PROJECT_DIR, 'pia_WPS_data.json')
    pia_xlsx = os.path.join(PROJECT_DIR, 'pia_WPS_data.xlsx')

    if os.path.exists(pia_json):
        print("=== Exporting PIA data ===")
        json_to_excel(pia_json, pia_xlsx, 'PIA')
    else:
        print("PIA JSON not found")

    # Export DFA data (if exists)
    dfa_json = os.path.join(PROJECT_DIR, 'dfa_WPS_data.json')
    dfa_xlsx = os.path.join(PROJECT_DIR, 'dfa_WPS_data.xlsx')

    if os.path.exists(dfa_json):
        print("\n=== Exporting DFA data ===")
        json_to_excel(dfa_json, dfa_xlsx, 'DFA')
    else:
        print("\nDFA JSON not found yet (DFA scraper may still be running)")

    # Merge both into combined file
    combined_xlsx = os.path.join(PROJECT_DIR, 'combined_WPS_data.xlsx')
    excel_files = [f for f in [pia_xlsx, dfa_xlsx] if os.path.exists(f)]

    if excel_files:
        print("\n=== Merging into combined file ===")
        merge_excels(excel_files, combined_xlsx)

    print("\nDone!")

if __name__ == '__main__':
    main()
